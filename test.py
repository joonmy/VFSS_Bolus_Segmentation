import argparse
import logging
import os
import random
import sys
import numpy as np
import torch
import torch.backends.cudnn as cudnn
import torch.nn as nn
from torch.utils.data import DataLoader
from tqdm import tqdm
import openpyxl as op
from openpyxl import Workbook
from datasets.dataset_vfss import VFSS_dataset, RandomGenerator_test
from utils import test_single_volume
from networks.vit_seg_modeling import VisionTransformer as ViT_seg
from networks.vit_seg_modeling import CONFIGS as CONFIGS_ViT_seg
from torchvision import transforms
os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
os.environ["CUDA_VISIBLE_DEVICES"] = "5,6"

parser = argparse.ArgumentParser()
parser.add_argument('--volume_path', type=str,
                    default='/home/harim/vfss_dt/Data/Dataset_Ver02', help='root dir for validation volume data')  # for acdc volume_path=root_dir
parser.add_argument('--dataset', type=str,
                    default='VFSS', help='experiment_name')
parser.add_argument('--num_classes', type=int,
                    default=6, help='output channel of network')
parser.add_argument('--max_iterations', type=int,
                    default=20000, help='maximum epoch number to train')
parser.add_argument('--max_epochs', type=int, default=250,
                    help='maximum epoch number to train')
parser.add_argument('--batch_size', type=int, default=16,
                    help='batch_size per gpu')
parser.add_argument('--img_size', type=int, default=224,
                    help='input patch size of network input')
parser.add_argument('--n_skip', type=int, default=3,
                    help='using number of skip-connect, default is num')
parser.add_argument('--vit_name', type=str,
                    default='R50-ViT-B_16', help='select one vit model')
parser.add_argument('--deterministic', type=int,  default=1,
                    help='whether use deterministic training')
parser.add_argument('--base_lr', type=float,  default=0.001,
                    help='segmentation network learning rate')
parser.add_argument('--seed', type=int, default=1234, help='random seed')
parser.add_argument('--vit_patches_size', type=int,
                    default=16, help='vit_patches_size, default is 16')
parser.add_argument('--test_case', type=str, default='test1', help='test case')
args = parser.parse_args()


def inference(args, model, test_save_path=None, grad_save_path=None):
    db_test = VFSS_dataset(base_dir=args.volume_path, split="test",
                           transform=transforms.Compose(
                               [RandomGenerator_test(output_size=[args.img_size, args.img_size])]))

    testloader = DataLoader(db_test, batch_size=1,
                            shuffle=False, num_workers=1)
    logging.info("{} test iterations per epoch".format(len(testloader)))
    model.eval()
    metric_list = 0.0
    metric_list2 = 0.0
    human_list = dict()
    human_list2 = dict()

    # Excel Creation
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('Each_images')
    write_ws = write_wb['Each_images']
    order = 1

    logit2_human_list = dict()
    logit2_human_list2 = dict()
    name = ['bolus', 'cervical spine', 'hyoid bone',
            'mandible', 'soft tissue', 'vocal folds']
    for i_batch, sampled_batch in tqdm(enumerate(testloader)):
        h, w = sampled_batch["image"].size()[2:]
        image, label, case_name = sampled_batch["image"], sampled_batch["label"], sampled_batch['case_name'][0]
        metric_i, metric_i2 = test_single_volume(image, label, model, classes=args.num_classes, patch_size=[args.img_size, args.img_size],
                                                 test_save_path=test_save_path, grad_save_path=grad_save_path, case=case_name, z_spacing=args.z_spacing)
        metric_list += np.array(metric_i)
        metric_list2 += np.array(metric_i2)
        logging.info('logit1: idx %d case %s mean_dice %f mean_hd95 %f \n Bolus : %f , Cervical : %f, Hyoid : %f, Mandible : %f, Soft tissue : %f, vocal : %f' % (
            i_batch, case_name, np.mean(metric_i, axis=0)[0], np.mean(metric_i, axis=0)[1], metric_i[0][0], metric_i[1][0], metric_i[2][0], metric_i[3][0], metric_i[4][0], metric_i[5][0]))

        logging.info('logit2: idx %d case %s mean_dice %f mean_hd95 %f \n Bolus : %f , Cervical : %f, Hyoid : %f, Mandible : %f, Soft tissue : %f, vocal : %f' % (
            i_batch, case_name, np.mean(metric_i2, axis=0)[0], np.mean(metric_i2, axis=0)[1], metric_i2[0][0], metric_i2[1][0], metric_i2[2][0], metric_i2[3][0], metric_i2[4][0], metric_i2[5][0]))

        # Write Excel
        write_ws['A'+str(order)] = case_name.split('/')[-1]

        write_ws['B'+str(order)] = 'bolus'
        write_ws['B'+str(order+1)] = 'cervical'
        write_ws['B'+str(order+2)] = 'hyoid'
        write_ws['B'+str(order+3)] = 'mandible'
        write_ws['B'+str(order+4)] = 'soft'
        write_ws['B'+str(order+5)] = 'vocal'
        write_ws['B'+str(order+6)] = 'total'

        write_ws['C'+str(order)] = format(metric_i[0][0], ".6f")
        write_ws['C'+str(order+1)] = format(metric_i[1][0], ".6f")
        write_ws['C'+str(order+2)] = format(metric_i[2][0], ".6f")
        write_ws['C'+str(order+3)] = format(metric_i[3][0], ".6f")
        write_ws['C'+str(order+4)] = format(metric_i[4][0], ".6f")
        write_ws['C'+str(order+5)] = format(metric_i[5][0], ".6f")
        write_ws['C'+str(order+6)] = format(metric_i[5][0], ".6f")

        write_ws['D'+str(order)] = format(metric_i2[0][0], ".6f")
        write_ws['D'+str(order+1)] = format(metric_i2[1][0], ".6f")
        write_ws['D'+str(order+2)] = format(metric_i2[2][0], ".6f")
        write_ws['D'+str(order+3)] = format(metric_i2[3][0], ".6f")
        write_ws['D'+str(order+4)] = format(metric_i2[4][0], ".6f")
        write_ws['D'+str(order+5)] = format(metric_i2[5][0], ".6f")
        write_ws['D'+str(order+6)] = format(metric_i2[5][0], ".6f")

        order = order + 8

        path = case_name.split("/")
        path = path[len(path)-1]
        path = path.split("_")[1]
        print(path)
        if path in human_list:
            human_list[path].append(np.mean(metric_i, axis=0)[0])
        else:
            human_list[path] = [np.mean(metric_i, axis=0)[0]]

        if path in human_list2:
            human_list2[path]["bolus"].append(metric_i[0][0])
            human_list2[path]["cervical"].append(metric_i[1][0])
            human_list2[path]["hyoid"].append(metric_i[2][0])
            human_list2[path]["mandible"].append(metric_i[3][0])
            human_list2[path]["soft"].append(metric_i[4][0])
            human_list2[path]["vocal"].append(metric_i[5][0])
        else:
            human_list2[path] = dict()
            human_list2[path]["bolus"] = [metric_i[0][0]]
            human_list2[path]["cervical"] = [metric_i[1][0]]
            human_list2[path]["hyoid"] = [metric_i[2][0]]
            human_list2[path]["mandible"] = [metric_i[3][0]]
            human_list2[path]["soft"] = [metric_i[4][0]]
            human_list2[path]["vocal"] = [metric_i[5][0]]

        if path in logit2_human_list:
            logit2_human_list[path].append(np.mean(metric_i2, axis=0)[0])
        else:
            logit2_human_list[path] = [np.mean(metric_i2, axis=0)[0]]

        if path in logit2_human_list2:
            logit2_human_list2[path]["bolus"].append(metric_i2[0][0])
            logit2_human_list2[path]["cervical"].append(metric_i2[1][0])
            logit2_human_list2[path]["hyoid"].append(metric_i2[2][0])
            logit2_human_list2[path]["mandible"].append(metric_i2[3][0])
            logit2_human_list2[path]["soft"].append(metric_i2[4][0])
            logit2_human_list2[path]["vocal"].append(metric_i2[5][0])
        else:
            logit2_human_list2[path] = dict()
            logit2_human_list2[path]["bolus"] = [metric_i2[0][0]]
            logit2_human_list2[path]["cervical"] = [metric_i2[1][0]]
            logit2_human_list2[path]["hyoid"] = [metric_i2[2][0]]
            logit2_human_list2[path]["mandible"] = [metric_i2[3][0]]
            logit2_human_list2[path]["soft"] = [metric_i2[4][0]]
            logit2_human_list2[path]["vocal"] = [metric_i2[5][0]]

    # Reset order and change sheet
    order = 1
    write_ws = write_wb.create_sheet('Individuals')
    write_ws = write_wb['Individuals']

    for key, value in human_list.items():
        print(key)

        mean = sum(human_list[key]) / len(human_list[key])
        mean_bolus = sum(human_list2[key]['bolus']) / \
            len(human_list2[key]['bolus'])
        mean_cervical = sum(
            human_list2[key]['cervical']) / len(human_list2[key]['cervical'])
        mean_hyoid = sum(human_list2[key]['hyoid']) / \
            len(human_list2[key]['hyoid'])
        mean_mandible = sum(
            human_list2[key]['mandible']) / len(human_list2[key]['mandible'])
        mean_soft = sum(human_list2[key]['soft']) / \
            len(human_list2[key]['soft'])
        mean_vocal = sum(human_list2[key]['vocal']) / \
            len(human_list2[key]['vocal'])

        logging.info('%s logit1: bolus : %s', key, mean_bolus)
        logging.info('%s logit1: cervical : %s', key, mean_cervical)
        logging.info('%s logit1: hyoid : %s', key, mean_hyoid)
        logging.info('%s logit1: mandible : %s', key, mean_mandible)
        logging.info('%s logit1: soft : %s', key, mean_soft)
        logging.info('%s logit1: vocal : %s', key, mean_vocal)
        logging.info('%s logit1: total: %s\n', key, mean)

        mean2 = sum(logit2_human_list[key]) / len(logit2_human_list[key])
        mean_bolus2 = sum(
            logit2_human_list2[key]['bolus']) / len(logit2_human_list2[key]['bolus'])
        mean_cervical2 = sum(
            logit2_human_list2[key]['cervical']) / len(logit2_human_list2[key]['cervical'])
        mean_hyoid2 = sum(
            logit2_human_list2[key]['hyoid']) / len(logit2_human_list2[key]['hyoid'])
        mean_mandible2 = sum(
            logit2_human_list2[key]['mandible']) / len(logit2_human_list2[key]['mandible'])
        mean_soft2 = sum(
            logit2_human_list2[key]['soft']) / len(logit2_human_list2[key]['soft'])
        mean_vocal2 = sum(
            logit2_human_list2[key]['vocal']) / len(logit2_human_list2[key]['vocal'])

        logging.info('%s logit2: bolus : %s', key, mean_bolus2)
        logging.info('%s logit2: cervical : %s', key, mean_cervical2)
        logging.info('%s logit2: hyoid : %s', key, mean_hyoid2)
        logging.info('%s logit2: mandible : %s', key, mean_mandible2)
        logging.info('%s logit2: soft : %s', key, mean_soft2)
        logging.info('%s logit2: vocal : %s', key, mean_vocal2)
        logging.info('%s logit2: total: %s\n', key, mean2)

        # Write Excel
        write_ws['A'+str(order)] = key

        write_ws['B'+str(order)] = 'bolus'
        write_ws['B'+str(order+1)] = 'cervical'
        write_ws['B'+str(order+2)] = 'hyoid'
        write_ws['B'+str(order+3)] = 'mandible'
        write_ws['B'+str(order+4)] = 'soft'
        write_ws['B'+str(order+5)] = 'vocal'
        write_ws['B'+str(order+6)] = 'total'

        write_ws['C'+str(order)] = format(mean_bolus, ".6f")
        write_ws['C'+str(order+1)] = format(mean_cervical, ".6f")
        write_ws['C'+str(order+2)] = format(mean_hyoid, ".6f")
        write_ws['C'+str(order+3)] = format(mean_mandible, ".6f")
        write_ws['C'+str(order+4)] = format(mean_soft, ".6f")
        write_ws['C'+str(order+5)] = format(mean_vocal, ".6f")
        write_ws['C'+str(order+6)] = format(mean, ".6f")

        write_ws['D'+str(order)] = format(mean_bolus2, ".6f")
        write_ws['D'+str(order+1)] = format(mean_cervical2, ".6f")
        write_ws['D'+str(order+2)] = format(mean_hyoid2, ".6f")
        write_ws['D'+str(order+3)] = format(mean_mandible2, ".6f")
        write_ws['D'+str(order+4)] = format(mean_soft2, ".6f")
        write_ws['D'+str(order+5)] = format(mean_vocal2, ".6f")
        write_ws['D'+str(order+6)] = format(mean2, ".6f")

        order = order + 8

    # Reset order and change sheet
    order = 1
    write_ws = write_wb.create_sheet('Total')
    write_ws = write_wb['Total']

    metric_list = metric_list / len(db_test)
    metric_list2 = metric_list2 / len(db_test)
    performance = np.mean(metric_list, axis=0)[0]
    mean_hd95 = np.mean(metric_list, axis=0)[1]
    logging.info('Logit1 Bolus: %f', metric_list[0][0])
    logging.info('Logit1 Cervical: %f', metric_list[1][0])
    logging.info('Logit1 hyoid: %f', metric_list[2][0])
    logging.info('Logit1 mandible: %f', metric_list[3][0])
    logging.info('Logit1 soft: %f', metric_list[4][0])
    logging.info('Logit1 vocal: %f', metric_list[5][0])
    logging.info('Testing performance of logit 1 in best val model: mean_dice : %f mean_hd95 : %f' % (
        performance, mean_hd95))
    performance2 = np.mean(metric_list2, axis=0)[0]
    mean_hd952 = np.mean(metric_list2, axis=0)[1]
    logging.info('Logit2 Bolus: %f', metric_list2[0][0])
    logging.info('Logit2 Cervical: %f', metric_list2[1][0])
    logging.info('Logit2 hyoid: %f', metric_list2[2][0])
    logging.info('Logit2 mandible: %f', metric_list2[3][0])
    logging.info('Logit2 soft: %f', metric_list2[4][0])
    logging.info('Logit2 vocal: %f', metric_list2[5][0])
    logging.info('Testing performance of logit 2 in best val model: mean_dice : %f mean_hd95 : %f' % (
        performance2, mean_hd952))

    # Write Excel
    write_ws['A'+str(order)] = 'Total Logits'

    write_ws['B'+str(order)] = 'bolus'
    write_ws['B'+str(order+1)] = 'cervical'
    write_ws['B'+str(order+2)] = 'hyoid'
    write_ws['B'+str(order+3)] = 'mandible'
    write_ws['B'+str(order+4)] = 'soft'
    write_ws['B'+str(order+5)] = 'vocal'
    write_ws['B'+str(order+6)] = 'total'

    write_ws['C'+str(order)] = format(metric_list[0][0], ".6f")
    write_ws['C'+str(order+1)] = format(metric_list[1][0], ".6f")
    write_ws['C'+str(order+2)] = format(metric_list[2][0], ".6f")
    write_ws['C'+str(order+3)] = format(metric_list[3][0], ".6f")
    write_ws['C'+str(order+4)] = format(metric_list[4][0], ".6f")
    write_ws['C'+str(order+5)] = format(metric_list[5][0], ".6f")
    write_ws['C'+str(order+6)] = format(performance, ".6f")

    write_ws['D'+str(order)] = format(metric_list2[0][0], ".6f")
    write_ws['D'+str(order+1)] = format(metric_list2[1][0], ".6f")
    write_ws['D'+str(order+2)] = format(metric_list2[2][0], ".6f")
    write_ws['D'+str(order+3)] = format(metric_list2[3][0], ".6f")
    write_ws['D'+str(order+4)] = format(metric_list2[4][0], ".6f")
    write_ws['D'+str(order+5)] = format(metric_list2[5][0], ".6f")
    write_ws['D'+str(order+6)] = format(performance2, ".6f")

    # Save Excel
    test_save_path = test_save_path.split('/')
    test_save_path = test_save_path[0] + "/" + \
        test_save_path[1] + "/" + test_save_path[2]
    print(test_save_path + "/" + args.test_case + "_result_best.xlsx")
    write_wb.save(test_save_path + "/" + args.test_case + "_result_best.xlsx")

    return "Testing Finished!"


if __name__ == "__main__":

    if not args.deterministic:
        cudnn.benchmark = True
        cudnn.deterministic = False
    else:
        cudnn.benchmark = False
        cudnn.deterministic = True
    random.seed(args.seed)
    np.random.seed(args.seed)
    torch.manual_seed(args.seed)
    torch.cuda.manual_seed(args.seed)

    dataset_config = {
        'VFSS': {
            'Dataset': VFSS_dataset,
            # 'volume_path': '/mnt/hdd3/VFSS/Data/Randomly_divided_dataset',
            'volume_path': "/home/younghun/VFSS/Data_0523Split[Undersampling]/3graychannel_0422",
            # 'volume_path': "/home/junmyeoung/VFSS/VFSS/Data/3graychannel_0422",
            'num_classes': 6,
            'z_spacing': 1,
        },
    }
    dataset_name = args.dataset
    args.num_classes = dataset_config[dataset_name]['num_classes']
    args.volume_path = dataset_config[dataset_name]['volume_path']
    args.Dataset = dataset_config[dataset_name]['Dataset']
    args.z_spacing = dataset_config[dataset_name]['z_spacing']
    args.is_pretrain = True

    # name the same snapshot defined in train script!
    #args.exp = 'TU_' + dataset_name + str(args.img_size)
    snapshot_path = "./Result/" + args.test_case
    # snapshot_path = "/home/harim/vfss_dt/Result_new_distribution/" + args.test_case
    # snapshot_path = snapshot_path + '_pretrain' if args.is_pretrain else snapshot_path
    # snapshot_path += '_' + args.vit_name
    # snapshot_path = snapshot_path + '_skip' + str(args.n_skip)
    # snapshot_path = snapshot_path + '_vitpatch' + str(args.vit_patches_size) if args.vit_patches_size!=16 else snapshot_path
    # snapshot_path = snapshot_path + '_epo' + str(args.max_epochs) if args.max_epochs != 30 else snapshot_path
    # if dataset_name == 'ACDC':  # using max_epoch instead of iteration to control training duration
    #     snapshot_path = snapshot_path + '_' + str(args.max_iterations)[0:2] + 'k' if args.max_iterations != 30000 else snapshot_path
    # snapshot_path = snapshot_path+'_bs'+str(args.batch_size)
    # snapshot_path = snapshot_path + '_lr' + str(args.base_lr) if args.base_lr != 0.01 else snapshot_path
    # snapshot_path = snapshot_path + '_'+str(args.img_size)
    # snapshot_path = snapshot_path + '_s'+str(args.seed) if args.seed!=1234 else snapshot_path

    config_vit = CONFIGS_ViT_seg[args.vit_name]
    config_vit.n_classes = args.num_classes
    config_vit.n_skip = args.n_skip
    config_vit.patches.size = (args.vit_patches_size, args.vit_patches_size)

    if args.vit_name.find('R50') != -1:
        config_vit.patches.grid = (int(
            args.img_size/args.vit_patches_size), int(args.img_size/args.vit_patches_size))

    net = ViT_seg(config_vit, img_size=args.img_size,
                  num_classes=config_vit.n_classes).cuda()

    snapshot = os.path.join(
        # snapshot_path, args.test_case + '_' + 'bestmodel_epoch:41iternum:3760' + '.pth')
        snapshot_path, args.test_case + '_bestmodel_epoch:53iternum:4860.pth')
    net.load_state_dict(torch.load(snapshot))
    snapshot_name = snapshot_path.split('/')[-1]

    log_testfile = './Result/'+args.test_case + \
        '/logtest_' + args.test_case + \
        '_bestmodel_epoch:53iternum:4860.txt'
    logging.basicConfig(filename=log_testfile, level=logging.INFO,
                        format='[%(asctime)s.%(msecs)03d] %(message)s', datefmt='%H:%M:%S')
    logging.getLogger().addHandler(logging.StreamHandler(sys.stdout))
    logging.info(str(args))
    logging.info(snapshot_name)

    test_save_path = './Result/'+args.test_case + '/prediction_best.pth'
    # grad_save_path = './Result/'+args.test_case + '/gradCAM_de_last'
    grad_save_path = './Result/'+args.test_case + \
        '/gradCAM_net.transformer.embeddings.hybrid_model.body.block3.unit9.relu'
    os.makedirs(test_save_path, exist_ok=True)
    os.makedirs(grad_save_path, exist_ok=True)

    inference(args, net, test_save_path, grad_save_path)
